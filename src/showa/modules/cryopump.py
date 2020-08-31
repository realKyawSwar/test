from openpyxl import load_workbook
import pandas as pd
from showa.modules import config
# import config


path = config.cryopump
# workbook object is created
# wb_obj = openpyxl.load_workbook(path)
wb_obj = load_workbook(path, read_only=True,
                       data_only=True, keep_links=False)


def getCryo():
    df = pd.DataFrame()
    dueCalc = config.cryoPer * 30
    try:
        sheet_obj = wb_obj['Eq summary']
        lineNo = []
        eqtType = []
        serviceDate = []
        chNo = []
        x = 1
        y = 34
        for j in range(y+1, y+357):
            line_cell = sheet_obj.cell(row=j, column=x)
            eqt_cell = sheet_obj.cell(row=j, column=x+2)
            ch_cell = sheet_obj.cell(row=j, column=x+1)
            date_cell = sheet_obj.cell(row=j, column=x+4)
            if ch_cell.value is None or eqt_cell.value == "TMP":
                pass
            elif line_cell.value == 206 or line_cell.value == 207 or line_cell.value is None:
                pass
            else:
                eqtType.append(eqt_cell.value)
                serviceDate.append(date_cell.value)
                lineNo.append(line_cell.value)
                chNo.append(ch_cell.value)
        zippedList = list(zip(lineNo, chNo, eqtType, serviceDate))
        df = pd.DataFrame(zippedList, columns=['Line', 'Location', 'Type',
                          'Date'])
        df['Date'] = pd.to_datetime(df['Date'])
        df['Due Date'] = df['Date'] + pd.DateOffset(days=dueCalc)
    except Exception as e:
        print("Cryopump error: {}".format(e))
    return df


# print(getCryo())
