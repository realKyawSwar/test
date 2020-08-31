from openpyxl import load_workbook
import pandas as pd
from concurrent.futures import ProcessPoolExecutor
from showa.modules import config
# import config


path = config.gvcylinder
# workbook object is created
wb_obj = load_workbook(path, read_only=True,
                       data_only=True, keep_links=False)


def getGV(line):
    sheet_obj = wb_obj['{}'.format(line)]
    lineNo = []
    eqtType = []
    serviceDate = []
    chNo = []
    x = 2
    y = 53
    for j in range(y+1, y+39):
        ch_cell = sheet_obj.cell(row=j, column=x+3)
        date_cell = sheet_obj.cell(row=j, column=x+2)
        if date_cell.value is None:
            pass
        elif line == 205:
            eqtType.append('PGV Cylinder')
            serviceDate.append(date_cell.value)
            lineNo.append(line)
            chNo.append(ch_cell.value)
        else:
            eqtType.append('EGV Cylinder')
            serviceDate.append(date_cell.value)
            lineNo.append(line)
            chNo.append(ch_cell.value)
    zippedList = list(zip(lineNo, chNo, eqtType, serviceDate))
    df = pd.DataFrame(zippedList, columns=['Line', 'Location', 'Type', 'Date'])
    return df


def getVATM(line):
    sheet_obj = wb_obj['{}'.format(line)]
    lineNo = []
    eqtType = []
    serviceDate = []
    chNo = []
    x = 2
    for j in range(94, 95):
        ch_cell = sheet_obj.cell(row=j, column=x+3)
        date_cell = sheet_obj.cell(row=j, column=x+2)
        if date_cell.value is None:
            pass
        else:
            eqtType.append('VAC Cylinder')
            serviceDate.append(date_cell.value)
            lineNo.append(line)
            chNo.append(ch_cell.value)
    for j in range(98, 99):
        ch_cell = sheet_obj.cell(row=j, column=x+3)
        date_cell = sheet_obj.cell(row=j, column=x+2)
        if date_cell.value is None:
            pass
        else:
            eqtType.append('ATM Cylinder')
            serviceDate.append(date_cell.value)
            lineNo.append(line)
            chNo.append(ch_cell.value)
    zippedList = list(zip(lineNo, chNo, eqtType, serviceDate))
    df = pd.DataFrame(zippedList, columns=['Line', 'Location', 'Type', 'Date'])
    return df


def allLinesGV():
    dueCalc = config.gvPer * 30
    lineList = [201, 202, 203, 204, 211, 208, 209, 210, 205]
    combined = pd.DataFrame(columns=['Line', 'Location', 'Type', 'Date'])
    with ProcessPoolExecutor() as pool:
        for ws in pool.map(getGV, lineList):
            combined = pd.concat([combined, ws])
        for i in pool.map(getVATM, lineList):
            combined = pd.concat([combined, i])
    combined.reset_index(drop=True)
    combined['Date'] = pd.to_datetime(combined['Date'])
    combined['Due Date'] = combined['Date'] + pd.DateOffset(days=dueCalc)
    return combined


if __name__ == "__main__":
    df = allLinesGV()
    df = df.reset_index(drop=True)
    print(df)


# # listy = df.Line.tolist()
# # for i in listy:
# #     print(i)
