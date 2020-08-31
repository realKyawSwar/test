from openpyxl import load_workbook
import datetime
from showa.modules import config
# import config
import pandas as pd

today = datetime.date.today()
print(today)
path = config.regen
wb_obj = load_workbook(path, read_only=True,
                       data_only=True, keep_links=False)
sheet = wb_obj.active


def getTodaycol():
    # searching
    # result = 0
    for value in sheet.iter_rows(min_row=4, max_row=4, min_col=470,
                                 values_only=True):
        for x, i in enumerate(value):
            print(i)
            if i is not None and i.date() == today:
                result = x+466
                print("x is {}".format(x))
                print(result)
                break
    return(result)


def getRegendate(line):
    y = getTodaycol()
    print(y)
    rowNo = 4*(int(line)-200)+1
    result = 0
    for value in sheet.iter_rows(min_row=rowNo, max_row=rowNo, min_col=y,
                                 values_only=True):
        for x, i in enumerate(value):
            if i == 'REGEN(AM)':
                result = x + 1
                break
            else:
                result = -1
    return result


def regenize():
    df = pd.DataFrame()
    try:
        lineNo = []
        daysToRe = []
        regenDate = []
        for i in range(201, 212):
            x = getRegendate(i)
            if i == 206 and i == 207:
                pass
            elif x < 0:
                pass
            else:
                lineNo.append(i)
                daysToRe.append(x)
                y = datetime.timedelta(x)+today
                regenDate.append(y)
        tempList = list(zip(lineNo, daysToRe, regenDate))
        df = pd.DataFrame(tempList, columns=['Line', 'Days', 'Regen Date'])
        df['Regen Date'] = pd.to_datetime(df['Regen Date'])
    except Exception as e:
        print("Regen error: {}".format(e))
    return df


# print("{} days to Regen".format(getRegendate()))
# for value in sheet.iter_rows(min_row=5, max_row=5, min_col=8, max_col=30,
#                              values_only=True):
#     for i in value:
#         print(i)
# try:
#     for i in range(201, 211):
#         print(getRegendate(i))
# except Exception as err:
#     print(err)

# print(getRegendate(209))
