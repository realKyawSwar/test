from openpyxl import load_workbook
# from concurrent.futures import ProcessPoolExecutor
from time import perf_counter
import pandas as pd
from showa.modules import config
# import config
import numpy as np

begin = perf_counter()
path = config.allEqt
# workbook object is created
wb_obj = load_workbook(path, read_only=True,
                       data_only=True, keep_links=False)


def orifice():
    df = pd.DataFrame()
    dueCalc = config.orificePer * 30
    try:
        sheet_obj = wb_obj['Mbox,orifice']
        lineList = []
        dateList = []
        chList = []
        eqtList = []
        for i in range(5, 16):
            cell_obj = sheet_obj.cell(row=i, column=5)
            a = sheet_obj.cell(row=i, column=2)
            if a.value == 206 or a.value == 207:
                pass
            else:
                dateList.append(cell_obj.value)
                lineList.append(a.value)
                chList.append('All')
                eqtList.append('Orifice')
        zippedList = list(zip(lineList, chList, eqtList, dateList))
        df = pd.DataFrame(zippedList, columns=['Line', 'Location', 'Type',
                          'Date'])
        df['Date'] = pd.to_datetime(df['Date'])
        df['Due Date'] = df['Date'] + pd.DateOffset(days=dueCalc)
    except Exception as e:
        print('Orifice Error: {}'.format(e))
    return df


def carrierForce():
    df = pd.DataFrame()
    dueCalc = config.forcePer * 30
    try:
        sheet_obj = wb_obj['Rail Alignment']
        lineList = []
        dateList = []
        chList = []
        eqtList = []
        for i in range(2, 13):
            cell_obj = sheet_obj.cell(row=i, column=4)
            a = sheet_obj.cell(row=i, column=2)
            if a.value == 206 or a.value == 207:
                pass
            else:
                dateList.append(cell_obj.value)
                lineList.append(a.value)
                chList.append('All')
                eqtList.append('Carrier Force')
        zippedList = list(zip(lineList, chList, eqtList, dateList))
        df = pd.DataFrame(zippedList, columns=['Line', 'Location', 'Type',
                          'Date'])
        df['Date'] = pd.to_datetime(df['Date'])
        df['Due Date'] = df['Date'] + pd.DateOffset(days=dueCalc)
    except Exception as e:
        print('Carrier Force Error: {}'.format(e))
    return df


def TDUgreasing():
    df = pd.DataFrame()
    dueCalc = config.tdugreasePer * 30
    try:
        sheet_obj = wb_obj['TDU greasing']
        lineNo = []
        serviceDate = []
        typeList = []
        chList = []
        # temp = {}
        for i in range(2, 13):
            cell_obj = sheet_obj.cell(row=i, column=3)
            a = sheet_obj.cell(row=i, column=2)
            if a.value == 206 or a.value == 207:
                pass
            else:
                serviceDate.append(cell_obj.value)
                lineNo.append(a.value)
                typeList.append('TDU Greasing')
                chList.append('ALL')
        df = pd.DataFrame(list(zip(lineNo, chList, typeList, serviceDate)),
                          columns=['Line', 'Location', 'Type', 'Date'])
        df['Date'] = pd.to_datetime(df['Date'])
        df['Due Date'] = df['Date'] + pd.DateOffset(days=dueCalc)
    except Exception as e:
        print('TDU Greasing Error: {}'.format(e))
    return df


def vacuumMotor():
    combDf = pd.DataFrame()
    dueCalc = config.vacmPer * 30
    try:
        sheet_obj = wb_obj['Vacuum motor']
        lineNo = []
        loadserviceDate = []
        unloadserviceDate = []
        rotserviceDate = []
        chList = []
        x = 2
        y = 2
        for a in range(x+1, x+4):
            cell_obj = sheet_obj.cell(row=y, column=a)
            chList.append(cell_obj.value)
        for j in range(y+1, y+12):
            line_obj = sheet_obj.cell(row=j, column=x)
            ldate_obj = sheet_obj.cell(row=j, column=x+1)
            uldate_obj = sheet_obj.cell(row=j, column=x+2)
            rotdate_obj = sheet_obj.cell(row=j, column=x+3)
            if line_obj.value == 206 or line_obj.value == 207:
                pass
            else:
                lineNo.append(line_obj.value)
                loadserviceDate.append(ldate_obj.value)
                unloadserviceDate.append(uldate_obj.value)
                rotserviceDate.append(rotdate_obj.value)
        loadDict = dict(zip(lineNo, loadserviceDate))
        loadList = []
        llineList = []
        lchList = []
        ldateList = []
        for lines, dates in zip(list(loadDict.keys()),
                                list(loadDict.values())):
            llineList.append(lines)
            lchList.append('Load')
            loadList.append('Vacuum Motor')
            ldateList.append(dates)
        loadDf = pd.DataFrame(list(zip(llineList, lchList, loadList,
                              ldateList)), columns=['Line', 'Location', 'Type',
                              'Date'])
        unloadDict = dict(zip(lineNo, unloadserviceDate))
        unloadList = []
        unllineList = []
        unlchList = []
        unldateList = []
        for lines, dates in zip(list(unloadDict.keys()),
                                list(unloadDict.values())):
            unllineList.append(lines)
            unlchList.append('Unload')
            unloadList.append('Vacuum Motor')
            unldateList.append(dates)
        unloadDf = pd.DataFrame(list(zip(unllineList, unlchList, unloadList,
                                unldateList)), columns=['Line', 'Location',
                                'Type', 'Date'])
        rotDict = dict(zip(lineNo, rotserviceDate))
        print(rotDict)
        rotList = []
        rotlineList = []
        rotchList = []
        rotdateList = []
        for lines, dates in zip(list(rotDict.keys()), list(rotDict.values())):
            if dates == '-':
                pass
            else:
                rotlineList.append(lines)
                rotchList.append('Rot')
                rotList.append('Vacuum Motor')
                rotdateList.append(dates)
        rotDf = pd.DataFrame(list(zip(rotlineList, rotchList, rotList,
                             rotdateList)), columns=['Line', 'Location',
                             'Type', 'Date'])
        combDf = pd.concat([loadDf, unloadDf, rotDf])
        combDf['Date'] = pd.to_datetime(combDf['Date'])
        combDf['Due Date'] = combDf['Date'] + pd.DateOffset(days=dueCalc)
    except Exception as e:
        print('Vacuum Motor Error: {}'.format(e))
    return combDf


def carrierLock():
    combDf = pd.DataFrame()
    dueCalc = config.lockPer * 30
    try:
        sheet_obj = wb_obj['carrier lock asembly']
        lineNo = []
        loadserviceDate = []
        unloadserviceDate = []
        rotserviceDate = []
        chList = []
        x = 2
        y = 4
        for a in range(x+1, x+4):
            cell_obj = sheet_obj.cell(row=y, column=a)
            chList.append(cell_obj.value)
        for j in range(y+1, y+12):
            line_obj = sheet_obj.cell(row=j, column=x)
            lc_obj = sheet_obj.cell(row=j, column=x+1)
            ulc_obj = sheet_obj.cell(row=j, column=x+2)
            rot_obj = sheet_obj.cell(row=j, column=x+3)
            if line_obj.value == 206 or line_obj.value == 207:
                pass
            else:
                lineNo.append(line_obj.value)
                loadserviceDate.append(lc_obj.value)
                unloadserviceDate.append(ulc_obj.value)
                rotserviceDate.append(rot_obj.value)
        loadDict = dict(zip(lineNo, loadserviceDate))
        loadList = []
        llineList = []
        lchList = []
        ldateList = []
        for lines, dates in zip(list(loadDict.keys()),
                                list(loadDict.values())):
            llineList.append(lines)
            lchList.append('Load')
            loadList.append('Carrier Lock')
            ldateList.append(dates)
        loadDf = pd.DataFrame(list(zip(llineList, lchList, loadList,
                              ldateList)), columns=['Line', 'Location', 'Type',
                              'Date'])
        unloadDict = dict(zip(lineNo, unloadserviceDate))
        unloadList = []
        unllineList = []
        unlchList = []
        unldateList = []
        for lines, dates in zip(list(unloadDict.keys()),
                                list(unloadDict.values())):
            unllineList.append(lines)
            unlchList.append('Unload')
            unloadList.append('Carrier Lock')
            unldateList.append(dates)
        unloadDf = pd.DataFrame(list(zip(unllineList, unlchList, unloadList,
                                unldateList)), columns=['Line', 'Location',
                                'Type', 'Date'])
        rotDict = dict(zip(lineNo, rotserviceDate))
        rotList = []
        rotlineList = []
        rotchList = []
        rotdateList = []
        for lines, dates in zip(list(rotDict.keys()), list(rotDict.values())):
            if dates == '-':
                continue
            else:
                rotlineList.append(lines)
                rotchList.append('Rot')
                rotList.append('Carrier Lock')
                rotdateList.append(dates)
        rotDf = pd.DataFrame(list(zip(rotlineList, rotchList, rotList,
                             rotdateList)), columns=['Line', 'Location',
                             'Type', 'Date'])
        combDf = pd.concat([loadDf, unloadDf, rotDf])
        combDf['Date'] = pd.to_datetime(combDf['Date'])
        combDf['Due Date'] = combDf['Date'] + pd.DateOffset(days=dueCalc)
    except Exception as e:
        print('Carrier Lock Error: {}'.format(e))
    return combDf


def bias():
    df = pd.DataFrame()
    dueCalc = config.biasPer * 30
    try:
        sheet_obj = wb_obj['Bias']
        lineNo = []
        eqtType = []
        serviceDate = []
        chNo = []
        x = 2
        y = 3
        for j in range(y+1, y+92):
            line_cell = sheet_obj.cell(row=j, column=x)
            eqt_cell = sheet_obj.cell(row=j, column=x+5)
            ch_cell = sheet_obj.cell(row=j, column=x+1)
            date_cell = sheet_obj.cell(row=j, column=x+3)
            if ch_cell.value is None:
                pass
            else:
                eqtType.append("{} Bias".format(eqt_cell.value))
                serviceDate.append(date_cell.value)
                lineNo.append(line_cell.value)
                chNo.append(ch_cell.value)
        zippedList = list(zip(lineNo, chNo, eqtType, serviceDate))
        df = pd.DataFrame(zippedList, columns=['Line', 'Location', 'Type',
                          'Date'])
        df['Date'] = pd.to_datetime(df['Date'])
        df['Due Date'] = df['Date'] + pd.DateOffset(days=dueCalc)
        df.dropna()
    except Exception as e:
        print('Bias Error: {}'.format(e))
    return df


def compressor():
    df = pd.DataFrame()
    dueCalc = config.compressPer * 30
    try:
        sheet_obj = wb_obj['Compressor1']
        lineNo = []
        eqtType = []
        serviceDate = []
        chNo = []
        x = 3
        y = 4
        for j in range(y+1, y+86):
            line_cell = sheet_obj.cell(row=j, column=x)
            eqt_cell = sheet_obj.cell(row=j, column=x+6)
            ch_cell = sheet_obj.cell(row=j, column=x+5)
            date_cell = sheet_obj.cell(row=j, column=x+2)
            if line_cell.value == 206 or line_cell.value == 207 or line_cell.value is None:
                pass
            else:
                eqtType.append(eqt_cell.value)
                serviceDate.append(date_cell.value)
                lineNo.append(line_cell.value)
                chNo.append("CRC{}".format(ch_cell.value))
        zippedList = list(zip(lineNo, chNo, eqtType, serviceDate))
        df = pd.DataFrame(zippedList, columns=['Line', 'Location', 'Type',
                          'Date'])
        df['Date'] = pd.to_datetime(df['Date'])
        df.dropna()
        df['Due Date'] = df['Date'] + pd.DateOffset(days=dueCalc)
    except Exception as e:
        print("Compressor error: {}".format(e))
    return df


def drypump():
    df = pd.DataFrame()
    dueCalc = config.mu600Per * 30
    try:
        sheet_obj = wb_obj['Dry pump']
        lineNo = []
        eqtType = []
        serviceDate = []
        chNo = []
        x = 2
        y = 3
        for j in range(y+1, y+130):
            line_cell = sheet_obj.cell(row=j, column=x)
            eqt_cell = sheet_obj.cell(row=j, column=x+4)
            ch_cell = sheet_obj.cell(row=j, column=x+5)
            date_cell = sheet_obj.cell(row=j, column=x+2)
            if eqt_cell.value == 'MU600' and type(line_cell.value) is int:
                if line_cell.value != 206 and line_cell.value != 207:
                    eqtType.append(eqt_cell.value)
                    serviceDate.append(date_cell.value)
                    lineNo.append(line_cell.value)
                    chNo.append(ch_cell.value)
            else:
                pass
        zippedList = list(zip(lineNo, chNo, eqtType, serviceDate))
        df = pd.DataFrame(zippedList, columns=['Line', 'Location', 'Type',
                          'Date'])
        df['Date'] = pd.to_datetime(df['Date'])
        df.dropna()
        df['Due Date'] = df['Date'] + pd.DateOffset(days=dueCalc)
    except Exception as e:
        print("Drypump error: {}".format(e))
    return df


def TDU():
    df = pd.DataFrame()
    dueCalc = config.tduPer * 30
    try:
        sheet_obj = wb_obj['TDU']
        lineNo = []
        eqtType = []
        serviceDate = []
        chNo = []
        x = 2
        y = 3
        for j in range(y+1, y+384):
            line_cell = sheet_obj.cell(row=j, column=x)
            # eqt_cell = sheet_obj.cell(row=j, column=x+4)
            ch_cell = sheet_obj.cell(row=j, column=x+1)
            date_cell = sheet_obj.cell(row=j, column=x+2)
            if ch_cell.value is None:
                pass
            elif line_cell.value == 206 or line_cell.value == 207:
                pass
            else:
                serviceDate.append(date_cell.value)
                lineNo.append(line_cell.value)
                chNo.append(ch_cell.value)
                # print()
                if ch_cell.value[:-1] == 'C':
                    eqtType.append("Corner TDU")
                else:
                    eqtType.append("Process TDU")
        zippedList = list(zip(lineNo, chNo, eqtType, serviceDate))
        df = pd.DataFrame(zippedList, columns=['Line', 'Location', 'Type',
                          'Date'])
        df['Date'] = pd.to_datetime(df['Date'])
        df['Due Date'] = df['Date'] + pd.DateOffset(days=dueCalc)
        df.dropna()
    except Exception as e:
        print("TDU error: {}".format(e))
    # print(df)
    return df


def robot():
    df = pd.DataFrame()
    ppRobotCalc = config.ppRobotPer * 30
    ppArmCalc = config.ppArmPer * 30
    xferRobotCalc = config.xferRobotPer * 30
    xferArmCalc = config.xferArmPer * 30
    epsonCalc = config.epsonPer * 30
    roboBattCalc = config.roboBatPer * 30
    try:
        sheet_obj = wb_obj['Robots']
        lineNo = []
        eqtType = []
        serviceDate = []
        chNo = []
        x = 2
        y = 7
        # listy = [j for j in range(y+1, y+121)]
        for j in range(y+1, y+121):
            line_cell = sheet_obj.cell(row=j, column=x)
            eqt_cell = sheet_obj.cell(row=j, column=x+3)
            ch_cell = sheet_obj.cell(row=j, column=x+4)
            date_cell = sheet_obj.cell(row=j, column=x+1)
            if line_cell.value == 206 or line_cell.value == 207:
                pass
            else:
                eqtType.append(eqt_cell.value)
                serviceDate.append(date_cell.value)
                lineNo.append(line_cell.value)
                chNo.append(ch_cell.value)
        zippedList = list(zip(lineNo, chNo, eqtType, serviceDate))
        df = pd.DataFrame(zippedList, columns=['Line', 'Location', 'Type',
                          'Date'])
        df['Date'] = pd.to_datetime(df['Date'])
        df['Due Date'] = np.where(df.Type.isin(['LD PP', 'ULD PP']), df['Date']
                                  + pd.DateOffset(days=ppRobotCalc), df.Date)
        df['Due Date'] = np.where(df.Type.isin(['LD PP Arm', 'ULD PP Arm']),
                                  df['Date'] + pd.DateOffset(days=ppArmCalc),
                                  df['Due Date'])
        df['Due Date'] = np.where(df.Type.isin(['LD Xfer Arm',
                                  'ULD Xfer Arm']),
                                  df['Date'] + pd.DateOffset(days=xferArmCalc),
                                  df['Due Date'])
        df['Due Date'] = np.where(df.Type.isin(['LD Xfer', 'ULD Xfer']),
                                  df['Date'] + pd.DateOffset
                                  (days=xferRobotCalc), df['Due Date'])
        df['Due Date'] = np.where(df.Type.isin(['LD Epson', 'ULD Epson']),
                                  df['Date'] + pd.DateOffset(days=epsonCalc),
                                  df['Due Date'])
        df['Due Date'] = np.where(df.Type == 'Robot Batteries', df['Date'] +
                                  pd.DateOffset(days=roboBattCalc),
                                  df['Due Date'])
    except Exception as e:
        print('Robot Error: {}'.format(e))
    return df


def getEqt():
    eqtdfList = [carrierLock(), orifice(), TDUgreasing(), vacuumMotor(),
                 bias(), drypump(), compressor(), TDU(), carrierForce(),
                 robot()]
    combined = pd.DataFrame(columns=['Line', 'Location', 'Type', 'Date',
                                     'Due Date'])
    for i in eqtdfList:
        combined = pd.concat([combined, i])
    combined.reset_index()
    return combined

# getEqt()
# df = robot()
# listy = df.Line.tolist()
# for i in listy:
#     print(i)
if __name__ == "__main__":
    print(getEqt())
    end = perf_counter()
    print("    Total time {0:.2f}s".format(end - begin))


# print(compressor())
# vacuumMotor()
